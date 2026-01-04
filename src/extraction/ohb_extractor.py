"""
Extractor para datos del Observatori de l'Habitatge de Barcelona (OHB).

Este módulo descarga y procesa archivos Excel del OHB que contienen
información sobre régimen de tenencia, concentración de propiedad,
y características del mercado de alquiler en Barcelona.
"""

import logging
import pandas as pd
import requests
from pathlib import Path
from typing import Optional, Dict, List
from datetime import datetime

logger = logging.getLogger(__name__)

class OHBExtractor:
    """Extractor para datos del Observatori de l'Habitatge de Barcelona."""
    
    # URLs de archivos Excel del OHB
    EXCEL_URLS = {
        "regimen_tenencia": "https://www.ohb.cat/wp-content/uploads/2023/04/1020_Llars_tinenca.xlsx",
        "autocontencio": "https://www.ohb.cat/wp-content/uploads/2023/04/1010_Taxa_autocontencio.xlsx",
        "tamano_propietario": "https://www.ohb.cat/wp-content/uploads/2023/04/2010_Habitatges_grandaria_propietari2-1.xlsx",
        "tipo_propietario": "https://www.ohb.cat/wp-content/uploads/2023/04/2020_Habitatges_tipus_propietari2.xlsx",
        "edificios_residenciales": "https://www.ohb.cat/wp-content/uploads/2024/11/2030_Edificis_residencials2.xlsx",
        "alquiler_tamano_propietario": "https://www.ohb.cat/wp-content/uploads/2023/04/2050_Habitatges_lloguer_grandaria_propietari2.xlsx",
        "alquiler_tipo_propietario": "https://www.ohb.cat/wp-content/uploads/2023/04/2060_Habitatges_lloguer_tipus_propietari2.xlsx",
    }
    
    def __init__(self, data_dir: Optional[Path] = None):
        """
        Inicializa el extractor de OHB.
        
        Args:
            data_dir: Directorio donde guardar los archivos descargados.
                     Por defecto: data/raw/ohb/
        """
        if data_dir is None:
            project_root = Path(__file__).parent.parent.parent
            data_dir = project_root / "data" / "raw" / "ohb"
        
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"OHBExtractor inicializado. Directorio de datos: {self.data_dir}")
    
    def download_excel(self, dataset_key: str, force: bool = False) -> Optional[Path]:
        """
        Descarga un archivo Excel del OHB.
        
        Args:
            dataset_key: Clave del dataset (ej: 'regimen_tenencia')
            force: Si True, descarga aunque el archivo ya exista
            
        Returns:
            Path al archivo descargado, o None si falla
        """
        if dataset_key not in self.EXCEL_URLS:
            logger.error(f"Dataset desconocido: {dataset_key}")
            logger.info(f"Datasets disponibles: {list(self.EXCEL_URLS.keys())}")
            return None
        
        url = self.EXCEL_URLS[dataset_key]
        filename = f"{dataset_key}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = self.data_dir / filename
        
        # Si ya existe y no forzamos descarga, retornar el existente
        if filepath.exists() and not force:
            logger.info(f"Archivo ya existe: {filepath}")
            return filepath
        
        try:
            logger.info(f"Descargando {dataset_key} desde {url}...")
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            
            with open(filepath, 'wb') as f:
                f.write(response.content)
            
            logger.info(f"✓ Descargado: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error descargando {dataset_key}: {e}")
            return None
    
    def extract_regimen_tenencia(self, filepath: Optional[Path] = None) -> pd.DataFrame:
        """
        Extrae datos de régimen de tenencia de la hoja 'Total'.
        """
        if filepath is None:
            filepath = self.download_excel("regimen_tenencia")
            if filepath is None: return pd.DataFrame()
        
        try:
            # Según docs/OHB_DATA_ANALYSIS.md, la hoja es 'Total'
            df = pd.read_excel(filepath, sheet_name="Total")
            
            # Limpiar nombres de columnas (eliminar espacios y saltos de línea)
            df.columns = [str(c).replace("\n", " ").strip() for c in df.columns]
            
            # Filtrar solo Barcelona (Codi_àmbit 080193) o AMB si se desea
            # Por ahora guardamos todo y filtramos en el load
            
            # Procesar años (2022/2023 -> anio_inicio=2022, anio_fin=2023)
            def parse_years(year_str):
                match = re.search(r'(\d{4})/(\d{4})', str(year_str))
                if match:
                    return int(match.group(1)), int(match.group(2))
                return None, None
                
            import re
            years = df['Any'].apply(parse_years)
            df['anio_inicio'] = years.apply(lambda x: x[0])
            df['anio_fin'] = years.apply(lambda x: x[1])
            
            # Mapeo de columnas según docs
            mapping = {
                'Àmbit': 'ambito',
                'Subtotal propietat(%)': 'propiedad_total',
                'Propietat totalment pagada(%)': 'propiedad_pagada',
                'Propietat amb pagament pendents(%)': 'propiedad_pendiente',
                'Subtotal lloguer(%)': 'alquiler_total',
                'Lloguer a preu de mercat(%)': 'alquiler_mercado',
                'Lloguer inferior a preu de mercat(%)': 'alquiler_social',
                'Cessió gratuïta(%)': 'cesion_gratuita'
            }
            
            # Renombrar solo las que existan
            df = df.rename(columns={k: v for k, v in mapping.items() if k in df.columns})
            
            # Columnas finales deseadas
            cols = ['ambito', 'anio_inicio', 'anio_fin', 'propiedad_total', 'propiedad_pagada', 
                    'propiedad_pendiente', 'alquiler_total', 'alquiler_mercado', 
                    'alquiler_social', 'cesion_gratuita']
            
            return df[[c for c in cols if c in df.columns]]
            
        except Exception as e:
            logger.error(f"Error procesando régimen de tenencia: {e}")
            return pd.DataFrame()

    def extract_concentracion_propiedad(self) -> pd.DataFrame:
        """
        Consolida datos de tipo y tamaño de propietario.
        """
        df_tipo = self.extract_tipo_propietario()
        df_tamano = self.extract_tamano_propietario()
        
        # TODO: Implementar merge inteligente cuando se conozcan las estructuras exactas
        # Por ahora retornamos una combinación básica o el más completo
        return df_tipo
    
    def extract_tamano_propietario(self, filepath: Optional[Path] = None) -> pd.DataFrame:
        """
        Extrae datos de viviendas por tamaño del propietario de la hoja 'Barcelona'.
        """
        if filepath is None:
            filepath = self.download_excel("tamano_propietario")
            if filepath is None: return pd.DataFrame()
        
        try:
            # Los archivos OHB suelen tener 'Barcelona' como hoja principal para la ciudad
            # Si falla, intentamos con la primera hoja
            try:
                df = pd.read_excel(filepath, sheet_name="Barcelona")
            except:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True)
                target_sheet = "Barcelona" if "Barcelona" in wb.sheetnames else wb.sheetnames[0]
                df = pd.read_excel(filepath, sheet_name=target_sheet)
            
            # Limpiar nombres de columnas
            df.columns = [str(c).replace("\n", " ").strip() for c in df.columns]
            
            # Buscar columnas de grandes tenedores (habitualmente >10 viviendas)
            gt_cols = [c for c in df.columns if any(x in c for x in ["Més de 10", "Més de 15", ">10", ">15"])]
            df['pct_grandes_tenedores'] = df[gt_cols].sum(axis=1) if gt_cols else 0
            
            # Extraer año si existe la columna 'Any' o similar
            year_col = next((c for c in df.columns if any(x in c.lower() for x in ["any", "año", "period"])), None)
            import re
            if year_col:
                def parse_year(y):
                    m = re.search(r'(\d{4})', str(y))
                    return int(m.group(1)) if m else None
                df['anio_inicio'] = df[year_col].apply(parse_year)
                df['anio_fin'] = df['anio_inicio']
            else:
                # Fallback: intentar extraer del nombre del archivo o usar año actual
                year_match = re.search(r'(\d{4})', filepath.name)
                df['anio_inicio'] = int(year_match.group(1)) if year_match else datetime.now().year
                df['anio_fin'] = df['anio_inicio']
            
            df['ambito'] = 'Barcelona'
            
            cols = ['ambito', 'anio_inicio', 'anio_fin', 'pct_grandes_tenedores']
            return df[[c for c in cols if c in df.columns]]
            
        except Exception as e:
            logger.error(f"Error procesando tamaño propietario: {e}")
            return pd.DataFrame()

    def extract_tipo_propietario(self, filepath: Optional[Path] = None) -> pd.DataFrame:
        """
        Extrae datos de viviendas por tipo de propietario (Persona física/jurídica).
        """
        if filepath is None:
            filepath = self.download_excel("tipo_propietario")
            if filepath is None: return pd.DataFrame()
        
        try:
            try:
                df = pd.read_excel(filepath, sheet_name="Barcelona")
            except:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True)
                target_sheet = "Barcelona" if "Barcelona" in wb.sheetnames else wb.sheetnames[0]
                df = pd.read_excel(filepath, sheet_name=target_sheet)
                
            df.columns = [str(c).replace("\n", " ").strip() for c in df.columns]
            
            # Mapeo típico: 'Persona física(%)', 'Persona jurídica(%)'
            mapping = {
                'Persona física(%)': 'pct_persona_fisica',
                'Persona jurídica(%)': 'pct_persona_juridica',
                'Persones físiques(%)': 'pct_persona_fisica',
                'Persones jurídiques(%)': 'pct_persona_juridica'
            }
            
            df = df.rename(columns={k: v for k, v in mapping.items() if k in df.columns})
            
            # Año
            year_col = next((c for c in df.columns if any(x in c.lower() for x in ["any", "año", "period"])), None)
            import re
            if year_col:
                df['anio_inicio'] = df[year_col].apply(lambda y: int(re.search(r'(\d{4})', str(y)).group(1)) if re.search(r'(\d{4})', str(y)) else None)
                df['anio_fin'] = df['anio_inicio']
            else:
                year_match = re.search(r'(\d{4})', filepath.name)
                df['anio_inicio'] = int(year_match.group(1)) if year_match else datetime.now().year
                df['anio_fin'] = df['anio_inicio']
                
            df['ambito'] = 'Barcelona'
            
            cols = ['ambito', 'anio_inicio', 'anio_fin', 'pct_persona_fisica', 'pct_persona_juridica']
            return df[[c for c in cols if c in df.columns]]
            
        except Exception as e:
            logger.error(f"Error procesando tipo propietario: {e}")
            return pd.DataFrame()
    
    def extract_all(self, datasets: Optional[List[str]] = None) -> Dict[str, pd.DataFrame]:
        """
        Extrae todos los datasets especificados (o todos si no se especifica).
        
        Args:
            datasets: Lista de claves de datasets a extraer. Si None, extrae todos.
            
        Returns:
            Diccionario con {dataset_key: DataFrame}
        """
        if datasets is None:
            datasets = list(self.EXCEL_URLS.keys())
        
        results = {}
        
        for dataset_key in datasets:
            logger.info(f"\n{'='*60}")
            logger.info(f"Extrayendo: {dataset_key}")
            logger.info(f"{'='*60}")
            
            # Descargar archivo
            filepath = self.download_excel(dataset_key)
            if filepath is None:
                logger.warning(f"Saltando {dataset_key} (descarga fallida)")
                continue
            
            # Extraer según tipo
            if dataset_key == "regimen_tenencia":
                df = self.extract_regimen_tenencia(filepath)
            elif dataset_key == "tamano_propietario":
                df = self.extract_tamano_propietario(filepath)
            elif dataset_key == "tipo_propietario":
                df = self.extract_tipo_propietario(filepath)
            else:
                # Para otros datasets, solo leer el Excel
                try:
                    df = pd.read_excel(filepath, sheet_name=0)
                    logger.info(f"Dataset genérico extraído: {len(df)} registros")
                except Exception as e:
                    logger.error(f"Error leyendo {dataset_key}: {e}")
                    df = pd.DataFrame()
            
            results[dataset_key] = df
        
        logger.info(f"\n{'='*60}")
        logger.info(f"Extracción completada: {len(results)} datasets")
        logger.info(f"{'='*60}")
        
        return results


def main():
    """Función principal para pruebas."""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    extractor = OHBExtractor()
    
    # Extraer datasets prioritarios
    priority_datasets = [
        "regimen_tenencia",
        "tamano_propietario",
        "tipo_propietario",
        "alquiler_tamano_propietario",
        "alquiler_tipo_propietario"
    ]
    
    results = extractor.extract_all(priority_datasets)
    
    # Mostrar resumen
    print("\n" + "="*60)
    print("RESUMEN DE EXTRACCIÓN")
    print("="*60)
    for dataset, df in results.items():
        print(f"\n{dataset}:")
        print(f"  Registros: {len(df)}")
        if not df.empty:
            print(f"  Columnas: {df.columns.tolist()}")
            print(f"  Muestra:")
            print(df.head(3))


if __name__ == "__main__":
    main()
